from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from movies.models import Movie


SILENT, NORMAL, VERBOSE, VERY_VERBOSE = 0, 1, 2, 3

class Command(BaseCommand):
    help = ("Imports movies from a local XLSX file. "
            "Expects title, release year, rating and rank.")

    def add_arguments(self, parser):
        # Positional arguments
        parser.add_argument("file_path",
                            nargs=1,
                            type=str)

    def handle(self, *args, **options):
        verbosity = options.get("verbosity", NORMAL)
        file_path = options["file_path"][0]

        wb = load_workbook(filename=file_path)
        ws = wb.worksheets[0]

        if verbosity >= NORMAL:
            self.stdout.write("=== Importing movies ===")

        index = 0
        rows = ws.iter_rows(min_row=2)  # skip the column captions
        for row in rows:
            index += 1
            row_values = [cell.value for cell in row]
            (title, release_year, rating, rank) = row_values
            movie, created = Movie.objects.get_or_create(
                title=title,
                release_year=release_year,
                rating=rating,
                rank=rank)
            if verbosity >= NORMAL:
                self.stdout.write(f"{movie.rank}. {movie.title}")
